import openpyxl as xl
from os import path
import smtplib, ssl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart


class GivenProduct:
    def __init__(self, makat, name, amount):
        self.makat = makat
        self.name = name
        self.amount = amount

    def __str__(self):
        return f"""פרטי המוצר הם,
מקט: {self.makat}
שם: {self.name}
כמות שהוקצתה:{self.amount}"""

    def __repr__(self):
        return f"""פרטי המוצר הם, מקט: {self.makat} שם: {self.name} כמות שהוקצתה: {self.amount}|"""


def send_mail_fin(person, product_list, reason):
    port = 465  # For SSL
    smtp_server = "smtp.gmail.com"
    sender_email = "dsfsd713@gmail.com"  # Enter your address
    receiver_email = f"{person}@gmail.com"  # Enter receiver address
    password = "123456ASD"  # Enter your password
    message = MIMEMultipart("alternative")
    message["Subject"] = f"הקצאת ציוד ל{person}"
    message["From"] = sender_email
    message["To"] = receiver_email

    # Create the plain-text and HTML version of your message
    html_prefix = f"""\
    <html>
<body>
 <p dir="rtl"><span style="font-family: 'Comic Sans MS', sans-serif; font-size: 14px;">היי,&nbsp;</span></p>
<p dir="rtl"><span style="font-family: 
        Comic Sans MS, sans-serif;"><span style="font-size: 
      14px;">מבקש להקצות ל{person} את הציוד הנ&quot;ל:</span></span></p>
<table style="width: 95%; float: right; margin-left: calc(5%);">
    <tbody>
        <tr style="height: 18.3333px;">
            <td style="width: 3.681%; background-color: rgb(250, 197, 28); text-align: center; height: 18.3333px;"><span style="font-size: 
      14px;"><strong><span style="font-family: 
        Comic Sans MS, sans-serif;">מס&quot;ד</span></strong></span></td>
            <td style="width: 21.7178%; background-color: rgb(250, 197, 28); height: 18.3333px;">
                <div style="text-align: center;"><span style="font-size: 
      14px;"><strong><span style="font-family: 
        Comic Sans MS, sans-serif;">שם המוצר</span></strong></span></div>
            </td>
            <td style="width: 25.5215%; background-color: rgb(250, 197, 28); height: 18.3333px;">
                <div style="text-align: center;"><span style="font-family: 
        Comic Sans MS, sans-serif;"><strong><span style="font-size: 
      14px;">מק&quot;ט</span></strong></span></div>
            </td>
            <td style="width: 8.957%; background-color: rgb(250, 197, 28); height: 18.3333px;">
                <div style="text-align: center;"><span style="font-family: 
        Comic Sans MS, sans-serif;"><strong><span style="font-size: 
      14px;">כמות</span></strong></span></div>
            </td>
            <td style="width: 19.3865%; background-color: rgb(250, 197, 28); height: 18.3333px;">
                <div data-empty="true" style="text-align: center;"><span style="font-family: 
        Comic Sans MS, sans-serif;"><strong><span style="font-size: 
      14px;">עבור</span></strong></span></div>
            </td>
            <td style="width: 22.0859%; background-color: rgb(250, 197, 28); height: 18.3333px;">
                <div data-empty="true" style="text-align: center;"><span style="font-family: 
        Comic Sans MS, sans-serif;"><strong><span style="font-size: 
      14px;">סיבה</span></strong></span></div>
            </td>
        </tr> """
    html_body = ""

    for idx, product_from_list in enumerate(product_list):
        html_body += f"""<tr style="height: 18px;">
            <td style="width: 3.681%; text-align: center; height: 18px;"><span style="font-family: 
        Comic Sans MS, sans-serif;"><span style="font-size: 14px;">{idx + 1}</span></span></td>
            <td style="width: 21.7178%; text-align: center; height: 18px;"><span style="font-family: 
        Comic Sans MS, sans-serif;"><span style="font-size: 
      14px;">{product_from_list.name}</span></span></td>
            <td style="width: 25.5215%; text-align: center; height: 18px;"><span style="font-family: 
        Comic Sans MS, sans-serif;"><span style="font-size: 
      14px;">{product_from_list.makat}<br></span></span></td>
            <td style="width: 8.957%; text-align: center; height: 18px;"><span style="font-family: 
        Comic Sans MS, sans-serif;"><span style="font-size: 
      14px;">{product_from_list.amount}</span></span></td>
            <td style="width: 19.3865%; text-align: center; height: 18px;"><span style="font-family: 
        Comic Sans MS, sans-serif;"><span style="font-size: 
      14px;">{person}</span></span></td>
            <td style="width: 22.0859%; text-align: center; height: 18px;"><span style="font-family: 
        Comic Sans MS, sans-serif;"><span style="font-size: 
      14px;">{reason}</span></span></td>
        </tr>"""

    html_fin = f""" </tbody>
</table>
<p dir="rtl"><span style="font-family: 
        Comic Sans MS, sans-serif;"><span style="font-size: 
      14px;">&nbsp;</span></span></p>
<p dir="rtl"><br></p>
<p dir="rtl"><span style="font-family: 
        Comic Sans MS, sans-serif;"><span style="font-size: 
      14px;"><br>שים לב ההקצאה תקפה לשבועיים. <br>{person}- לא לשכוח לתאם מועד משיכה מול זילכה.&nbsp;</span></span></p>
<p dir="rtl"><span style="font-family: 
        Comic Sans MS, sans-serif;"><span style="font-size: 
      14px;">בברכה,</span></span></p>
<p dir="rtl"><span style="font-family: 
        Comic Sans MS, sans-serif;"><span style="font-size: 
      14px;">עילם קדן</span></span></p>
<p><span style="font-family: 
        Comic Sans MS, sans-serif;"><span style="font-size: 
      14px;">&nbsp;</span></span></p>
<p><span style='font-size: 14px; font-family: "Comic Sans MS", sans-serif;'>&nbsp;</span></p>
</body>
</html>
    """
    html_full = html_prefix + html_body + html_fin
    # Turn these into plain/html MIMEText objects
    part2 = MIMEText(html_full, "html")
    # Add HTML/plain-text parts to MIMEMultipart message
    # The email client will try to render the last part first
    message.attach(part2)

    # Create secure connection with server and send email
    context = ssl.create_default_context()
    with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
        server.login(sender_email, password)
        server.sendmail(
            sender_email, receiver_email, message.as_string()
        )


def send_mail_fin_2(person_list, product_list, reason):
    person = person_list[0]
    port = 465  # For SSL
    smtp_server = "smtp.gmail.com"
    sender_email = "dsfsd713@gmail.com"  # Enter your address
    receiver_email = []
    receiver_email.append(f"{person}@gmail.com")
    send_to = f"{person}@gmail.com"
    receiver_email.append(f"{person}@gmail.com")  # Enter receiver address
    person_list.remove(person_list[0])
    for person_from_list in person_list:
        send_to = send_to + f", {person_from_list}@gmail.com"
        receiver_email.append(f"{person_from_list}@gmail.com")
    password = "123456ASD"
    message = MIMEMultipart("alternative")
    message["Subject"] = f"הקצאת ציוד ל{person}"
    message["From"] = sender_email
    message["To"] = send_to

    # Create the plain-text and HTML version of your message
    html_prefix = f"""\
    <html>
<body>
 <p dir="rtl"><span style="font-family: 'Comic Sans MS', sans-serif; font-size: 14px;">היי,&nbsp;</span></p>
<p dir="rtl"><span style="font-family: 
        Comic Sans MS, sans-serif;"><span style="font-size: 
      14px;">מבקש להקצות ל{person} את הציוד הנ&quot;ל:</span></span></p>
<table style="width: 95%; float: right; margin-left: calc(5%);">
    <tbody>
        <tr style="height: 18.3333px;">
            <td style="width: 3.681%; background-color: rgb(250, 197, 28); text-align: center; height: 18.3333px;"><span style="font-size: 
      14px;"><strong><span style="font-family: 
        Comic Sans MS, sans-serif;">מס&quot;ד</span></strong></span></td>
            <td style="width: 21.7178%; background-color: rgb(250, 197, 28); height: 18.3333px;">
                <div style="text-align: center;"><span style="font-size: 
      14px;"><strong><span style="font-family: 
        Comic Sans MS, sans-serif;">שם המוצר</span></strong></span></div>
            </td>
            <td style="width: 25.5215%; background-color: rgb(250, 197, 28); height: 18.3333px;">
                <div style="text-align: center;"><span style="font-family: 
        Comic Sans MS, sans-serif;"><strong><span style="font-size: 
      14px;">מק&quot;ט</span></strong></span></div>
            </td>
            <td style="width: 8.957%; background-color: rgb(250, 197, 28); height: 18.3333px;">
                <div style="text-align: center;"><span style="font-family: 
        Comic Sans MS, sans-serif;"><strong><span style="font-size: 
      14px;">כמות</span></strong></span></div>
            </td>
            <td style="width: 19.3865%; background-color: rgb(250, 197, 28); height: 18.3333px;">
                <div data-empty="true" style="text-align: center;"><span style="font-family: 
        Comic Sans MS, sans-serif;"><strong><span style="font-size: 
      14px;">עבור</span></strong></span></div>
            </td>
            <td style="width: 22.0859%; background-color: rgb(250, 197, 28); height: 18.3333px;">
                <div data-empty="true" style="text-align: center;"><span style="font-family: 
        Comic Sans MS, sans-serif;"><strong><span style="font-size: 
      14px;">סיבה</span></strong></span></div>
            </td>
        </tr> """
    html_body = ""

    for idx, product_from_list in enumerate(product_list):
        html_body = html_body + f"""<tr style="height: 18px;">
            <td style="width: 3.681%; text-align: center; height: 18px;"><span style="font-family: 
        Comic Sans MS, sans-serif;"><span style="font-size: 14px;">{idx + 1}</span></span></td>
            <td style="width: 21.7178%; text-align: center; height: 18px;"><span style="font-family: 
        Comic Sans MS, sans-serif;"><span style="font-size: 
      14px;">{product_from_list.name}</span></span></td>
            <td style="width: 25.5215%; text-align: center; height: 18px;"><span style="font-family: 
        Comic Sans MS, sans-serif;"><span style="font-size: 
      14px;">{product_from_list.makat}<br></span></span></td>
            <td style="width: 8.957%; text-align: center; height: 18px;"><span style="font-family: 
        Comic Sans MS, sans-serif;"><span style="font-size: 
      14px;">{product_from_list.amount}</span></span></td>
            <td style="width: 19.3865%; text-align: center; height: 18px;"><span style="font-family: 
        Comic Sans MS, sans-serif;"><span style="font-size: 
      14px;">{person}</span></span></td>
            <td style="width: 22.0859%; text-align: center; height: 18px;"><span style="font-family: 
        Comic Sans MS, sans-serif;"><span style="font-size: 
      14px;">{reason}</span></span></td>
        </tr>"""

    html_fin = f""" </tbody>
</table>
<p dir="rtl"><span style="font-family: 
        Comic Sans MS, sans-serif;"><span style="font-size: 
      14px;">&nbsp;</span></span></p>
<p dir="rtl"><br></p>
<p dir="rtl"><span style="font-family: 
        Comic Sans MS, sans-serif;"><span style="font-size: 
      14px;"><br>שים לב ההקצאה תקפה לשבועיים. <br>{person}- לא לשכוח לתאם מועד משיכה מול זילכה.&nbsp;</span></span></p>
<p dir="rtl"><span style="font-family: 
        Comic Sans MS, sans-serif;"><span style="font-size: 
      14px;">בברכה,</span></span></p>
<p dir="rtl"><span style="font-family: 
        Comic Sans MS, sans-serif;"><span style="font-size: 
      14px;">עילם קדן, &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; <strong>סגן</strong><br>קמ&quot;ד אינטגרציה ותקשורת<br>תחום&nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp;מיוחדים<br>&nbsp;</span></span></p>
<p><span style="font-family: 
        Comic Sans MS, sans-serif;"><span style="font-size: 
      14px;">&nbsp;</span></span></p>
<p><span style='font-size: 14px; font-family: "Comic Sans MS", sans-serif;'>&nbsp;</span></p>
</body>
</html>
    """
    html_full = html_prefix + html_body + html_fin
    # Turn these into plain/html MIMEText objects
    part2 = MIMEText(html_full, "html")
    # Add HTML/plain-text parts to MIMEMultipart message
    # The email client will try to render the last part first
    message.attach(part2)

    # Create secure connection with server and send email
    context = ssl.create_default_context()
    with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
        server.login(sender_email, password)
        server.sendmail(
            sender_email, receiver_email, message.as_string()
        )


def num_of_row(self, cell_test):
    for cell in self['B']:
        if cell.value == cell_test.value:
            return cell.row


def update_value(self, user_input):
    if int(self.value) - int(user_input) >= 0:
        self.value = int(self.value) - int(user_input)
    else:
        print("סליחה אחשלי אי אפשר")


def giving_procedure_2(makat, amount, given_products_list):
    cell_test = sheet['X1']
    cell_test.value = int(makat)
    if int(amount) < 0:
        print("fuck you you cant add minus")
    else:
        row_number = num_of_row(sheet, cell_test)
        if row_number != None:
            update_value(sheet[f'C{row_number}'], amount)
            print(
                f"""fyi you have {sheet[f'C{row_number}'].value} left from {sheet[f'A{row_number}'].value} that makat is {sheet[f'B{row_number}'].value}""")
            this_product = GivenProduct(makat, sheet[f'A{row_number}'].value, amount)
            given_products_list.append(this_product)
        else:
            print("אין מקט כזה גבר")


def giving_procedure_1(makat, amount, given_products_list, count_people):
    cell_test = sheet['X1']
    cell_test.value = int(makat)
    if int(amount) < 0:
        print("fuck you you cant add minus")
    else:
        row_number = num_of_row(sheet, cell_test)
        if row_number != None:
            update_value(sheet[f'C{row_number}'], int(amount) * int(count_people))
            print(
                f"""fyi you have {sheet[f'C{row_number}'].value} left from {sheet[f'A{row_number}'].value} that makat is {sheet[f'B{row_number}'].value}""")
            this_product = GivenProduct(makat, sheet[f'A{row_number}'].value, int(amount))
            given_products_list.append(this_product)
        else:
            print("אין מקט כזה גבר")


def load_workbook(wb_path):
    if path.exists(wb_path):
        return xl.load_workbook(wb_path)
    return xl.Workbook()


wb_path = 'transactions.xlsx'  # Enter your excel database here
wb = load_workbook(wb_path)
sheet = wb['1']
print(f"""Hi!
welcome to Eylam's program
you have several options:
you can send the same allocation to many recipients
you can send multiple cloned allocation to many recipients
you can add as many products as you want
you choose products by writing their makat number
***when you done adding the just write 'fin'***
""")
list_of_people = []
given_products_list = []
idx = 1
person = input("to whom you want to send the items? ")
list_of_people.append(person)
while person.upper() != "FIN":
    person = input("to whom you want to send the items? ")
    list_of_people.append(person)
list_of_people.remove(list_of_people[-1])
reason = input("and why?")
answer = input(f"""Choose your will
1) do you want to send one allocation for each one of them"
2) to send this allocation to all of them together
ENTER YOUR ANSWER 1 / 2""")
print(f"Product number {idx}")
makat = input("מה המקט?")
cell_test = sheet['X1']
cell_test.value = int(makat)
amount = input("how much?")
wb.save('transactions.xlsx')
if answer == "1":
    giving_procedure_1(makat, amount, given_products_list, len(list_of_people))
    while makat.upper() != "FIN" or amount.upper() != "FIN":
        idx = idx + 1
        print(f"Product number {idx}")
        makat = input("מה המקט?")
        if makat.upper() == "FIN":
            break
        cell_test = sheet['X1']
        cell_test.value = int(makat)
        amount = input("how much?")
        giving_procedure_1(makat, amount, given_products_list, len(list_of_people))
        wb.save('transactions.xlsx')

    for person_from_list in list_of_people:
        send_mail_fin(person_from_list, given_products_list, reason)

if answer == "2":
    giving_procedure_2(makat, amount, given_products_list)
    while makat.upper() != "FIN" or amount.upper() != "FIN":
        idx = idx + 1
        print(f"Product number {idx}")
        makat = input("מה המקט?")
        if makat.upper() == "FIN":
            break
        cell_test = sheet['X1']
        cell_test.value = int(makat)
        amount = input("how much?")
        giving_procedure_2(makat, amount, given_products_list)
        wb.save('transactions.xlsx')
    send_mail_fin_2(list_of_people, given_products_list, reason)
else:
    print("you should write 1 / 2")